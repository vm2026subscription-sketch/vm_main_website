"""
News routes for fetching and filtering RSS feeds
"""
import re
from datetime import datetime, timedelta
from email.utils import parsedate_to_datetime
from threading import Lock
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from xml.etree import ElementTree as ET

try:
    import feedparser
except ImportError:
    feedparser = None

from flask import Blueprint, jsonify, render_template, request, send_file

# Category keywords for filtering
CATEGORIES = {
    "entrance": {
        "keywords": [
            "jee", "neet", "mht-cet", "cuet", "gate", "cat", "clat", "upsc", "nda",
            "entrance exam", "admit card", "hall ticket", "answer key", "exam date",
            "question paper", "response sheet", "syllabus"
        ],
        "label": "Entrance Exams"
    },
    "results": {
        "keywords": [
            "result", "merit list", "scorecard", "rank list", "cut off", "declared",
            "cbse result", "toppers", "pass percentage"
        ],
        "label": "Results"
    },
    "admissions": {
        "keywords": [
            "admission", "counselling", "counseling", "allotment", "seat allotment", "cap round",
            "dte", "fyjc", "option form", "registration", "application form", "apply now"
        ],
        "label": "Admissions"
    },
    "govtjobs": {
        "keywords": ["recruitment", "vacancy", "govt job", "sarkari", "railway", "bank job", "notification"],
        "label": "Govt Jobs"
    },
    "scholarship": {
        "keywords": ["scholarship", "fellowship", "stipend", "financial aid"],
        "label": "Scholarships"
    }
}

# RSS Feed sources
RSS_FEEDS = [
    {"url": "https://admission.aglasem.com/feed", "source": "AglaSem"},
    {"url": "https://www.hindustantimes.com/feeds/rss/education/rssfeed.xml", "source": "Hindustan Times Education"},
    {"url": "https://www.indiatoday.in/rss/1206577", "source": "India Today Education"},
    {"url": "https://timesofindia.indiatimes.com/rssfeeds/913168846.cms", "source": "TOI Education"},
]

news_bp = Blueprint("news", __name__)

# Cache variables
_NEWS_CACHE = {}
_NEWS_CACHE_LOCK = Lock()
NEWS_CACHE_TTL_SECONDS = 600  # 10 minutes
DEFAULT_FRESHNESS_HOURS = 24


def _to_iso_datetime(raw_value):
    """Convert feed date values to naive ISO format used by this API."""
    if not raw_value:
        return datetime.now().isoformat()

    raw_value = str(raw_value).strip()
    if not raw_value:
        return datetime.now().isoformat()

    try:
        return parsedate_to_datetime(raw_value).replace(tzinfo=None).isoformat()
    except Exception:
        pass

    try:
        normalized = raw_value.replace("Z", "+00:00")
        return datetime.fromisoformat(normalized).replace(tzinfo=None).isoformat()
    except Exception:
        return datetime.now().isoformat()


_IMG_TAG_RE = re.compile(r'<img[^>]+src=["\']([^"\']+)["\']', re.IGNORECASE)


def _extract_img_from_html(html):
    """Return first <img src=...> URL found in HTML, or empty string."""
    m = _IMG_TAG_RE.search(html or "")
    return m.group(1) if m else ""


def _build_article(title, link, description, pub_date, source_name, image_url=""):
    """Normalize and filter a single feed entry to API article shape."""
    title = (title or "").strip()
    link = (link or "").strip()

    # Extract image from description HTML before stripping tags
    if not image_url:
        image_url = _extract_img_from_html(description)

    description = re.sub(r'<[^>]+>', '', (description or "")).strip()

    if not title or not link:
        return None

    category = _get_news_category(title, description)
    if not category:
        return None

    if len(description) > 200:
        description = description[:197] + "..."

    return {
        "title": title,
        "link": link,
        "date": _to_iso_datetime(pub_date),
        "desc": description,
        "source": source_name,
        "category": category,
        "image": image_url,
    }


def _fetch_feed_with_stdlib(feed_url, source_name, timeout=10):
    """Fallback RSS/Atom parser using Python stdlib when feedparser is unavailable."""
    articles = []
    try:
        req = Request(
            feed_url,
            headers={
                "User-Agent": "vm-main-website/1.0",
                "Accept": "application/rss+xml, application/xml, text/xml",
            },
        )
        with urlopen(req, timeout=timeout) as response:
            content = response.read()

        root = ET.fromstring(content)

        media_ns = 'http://search.yahoo.com/mrss/'

        def _item_image(el):
            """Extract best image URL from an RSS item/Atom entry element."""
            # <media:thumbnail url="...">
            thumb = el.find(f'{{{media_ns}}}thumbnail')
            if thumb is not None and thumb.get('url'):
                return thumb.get('url')
            # <media:content url="..." medium="image">
            for mc in el.findall(f'{{{media_ns}}}content'):
                if mc.get('url') and mc.get('medium', 'image') == 'image':
                    return mc.get('url')
            # <enclosure url="..." type="image/...">
            enc = el.find('enclosure')
            if enc is not None and (enc.get('type', '').startswith('image/') or not enc.get('type')):
                url = enc.get('url', '')
                if url:
                    return url
            return ''

        # RSS items
        for item in root.findall('.//item')[:50]:
            title = item.findtext('title', default='')
            link = item.findtext('link', default='')
            description = item.findtext('description', default='')
            pub_date = item.findtext('pubDate', default='')
            image_url = _item_image(item)
            article = _build_article(title, link, description, pub_date, source_name, image_url)
            if article:
                articles.append(article)

        # Atom entries (in case feed uses atom format)
        if not articles:
            atom_ns = {'atom': 'http://www.w3.org/2005/Atom'}
            for entry in root.findall('.//atom:entry', atom_ns)[:50]:
                title = entry.findtext('atom:title', default='', namespaces=atom_ns)
                link_elem = entry.find("atom:link[@rel='alternate']", atom_ns) or entry.find('atom:link', atom_ns)
                link = '' if link_elem is None else (link_elem.get('href') or '')
                description = entry.findtext('atom:summary', default='', namespaces=atom_ns) or entry.findtext('atom:content', default='', namespaces=atom_ns)
                pub_date = entry.findtext('atom:published', default='', namespaces=atom_ns) or entry.findtext('atom:updated', default='', namespaces=atom_ns)
                image_url = _item_image(entry)
                article = _build_article(title, link, description, pub_date, source_name, image_url)
                if article:
                    articles.append(article)

    except Exception as e:
        print(f"Error fetching feed {source_name}: {e}")
        return []

    return articles

def _get_news_category(title, description):
    """
    Determine news category based on keyword matching.
    Returns category key or None if no match.
    """
    text = (title + " " + description).lower()
    
    for category_key, category_data in CATEGORIES.items():
        for keyword in category_data["keywords"]:
            # Use word boundaries for better matching
            pattern = r'\b' + re.escape(keyword) + r'\b'
            if re.search(pattern, text):
                return category_key
    
    return None


def _fetch_feed(feed_url, source_name, timeout=10):
    """Fetch and parse RSS feed, return list of articles"""
    if feedparser is None:
        return _fetch_feed_with_stdlib(feed_url, source_name, timeout=timeout)
    
    articles = []
    try:
        feed_data = feedparser.parse(feed_url)
        
        for entry in feed_data.entries[:50]:  # Limit to 50 per feed
            title = entry.get("title", "")
            link = entry.get("link", "")
            description = entry.get("summary", entry.get("description", ""))
            pub_date = entry.get("published", entry.get("updated", ""))
            if hasattr(entry, "published_parsed") and entry.published_parsed:
                try:
                    pub_date = datetime(*entry.published_parsed[:6]).isoformat()
                except Exception:
                    pass

            # Extract image URL
            image_url = ""
            if hasattr(entry, "media_thumbnail") and entry.media_thumbnail:
                image_url = entry.media_thumbnail[0].get("url", "")
            elif hasattr(entry, "media_content") and entry.media_content:
                for mc in entry.media_content:
                    if mc.get("url") and mc.get("medium", "image") == "image":
                        image_url = mc["url"]
                        break
            elif hasattr(entry, "enclosures") and entry.enclosures:
                for enc in entry.enclosures:
                    if enc.get("type", "").startswith("image/"):
                        image_url = enc.get("href", enc.get("url", ""))
                        break

            article = _build_article(title, link, description, pub_date, source_name, image_url)
            if article:
                articles.append(article)
    
    except Exception as e:
        print(f"Error fetching feed {source_name}: {e}")
        return []
    
    return articles


def _get_all_news():
    """Fetch news from all feeds with caching"""
    with _NEWS_CACHE_LOCK:
        # Check if cache is still valid
        if _NEWS_CACHE.get("data") and _NEWS_CACHE.get("expires_at"):
            from datetime import datetime as dt
            if dt.now().timestamp() < _NEWS_CACHE["expires_at"]:
                return _NEWS_CACHE["data"]
    
    # Fetch fresh news from all feeds
    all_news = []
    for feed_info in RSS_FEEDS:
        articles = _fetch_feed(feed_info["url"], feed_info["source"])
        all_news.extend(articles)
    
    # Sort by date descending
    all_news.sort(key=lambda x: x["date"], reverse=True)
    
    # Update cache
    with _NEWS_CACHE_LOCK:
        from datetime import datetime as dt
        _NEWS_CACHE["data"] = all_news
        _NEWS_CACHE["expires_at"] = dt.now().timestamp() + NEWS_CACHE_TTL_SECONDS
    
    return all_news


def _parse_article_date(date_str):
    """Parse article date safely; fallback to minimum datetime on parse failure."""
    if not date_str:
        return datetime.min
    try:
        normalized = date_str.replace("Z", "+00:00")
        return datetime.fromisoformat(normalized).replace(tzinfo=None)
    except ValueError:
        return datetime.min


def _prioritize_freshness_by_category(articles, fresh_hours=DEFAULT_FRESHNESS_HOURS):
    """
    Prioritize recent items by category.
    - Include all recent (within fresh_hours) per category first.
    - If a category has no recent item, include one latest fallback from that category.
    - Append remaining older items afterwards.
    """
    if not articles:
        return []

    now = datetime.now()
    threshold = now - timedelta(hours=max(1, fresh_hours))

    parsed_items = []
    for article in articles:
        parsed_items.append((article, _parse_article_date(article.get("date"))))

    buckets = {
        cat: {"recent": [], "older": []}
        for cat in CATEGORIES.keys()
    }

    for article, article_dt in parsed_items:
        category = article.get("category")
        if category not in buckets:
            continue
        if article_dt >= threshold:
            buckets[category]["recent"].append((article, article_dt))
        else:
            buckets[category]["older"].append((article, article_dt))

    for category in buckets.keys():
        buckets[category]["recent"].sort(key=lambda item: item[1], reverse=True)
        buckets[category]["older"].sort(key=lambda item: item[1], reverse=True)

    ordered = []
    seen_links = set()

    def append_unique(article_obj):
        link_key = article_obj.get("link") or article_obj.get("title")
        if link_key in seen_links:
            return
        seen_links.add(link_key)
        ordered.append(article_obj)

    for category in CATEGORIES.keys():
        recent_items = buckets[category]["recent"]
        older_items = buckets[category]["older"]

        if recent_items:
            for article, _ in recent_items:
                append_unique(article)
        elif older_items:
            append_unique(older_items[0][0])

    remaining = [
        (article, article_dt)
        for article, article_dt in parsed_items
        if (article.get("link") or article.get("title")) not in seen_links
    ]
    remaining.sort(key=lambda item: item[1], reverse=True)
    for article, _ in remaining:
        append_unique(article)

    return ordered


@news_bp.route("/news-reel")
def news_reel():
    return render_template("news_reel.html")


@news_bp.route("/api/news", methods=["GET"])
def get_news():
    """
    Fetch filtered news from RSS feeds.
    Query params:
      - category: Filter by category (entrance, results, admissions, govtjobs, scholarship, all)
      - limit: Max number of articles (default 50)
            - fresh_hours: Recent-window in hours for all-category prioritization (default 24)
    """
    category_filter = request.args.get("category", "all").lower()
    limit = int(request.args.get("limit", 50))
    fresh_hours = int(request.args.get("fresh_hours", DEFAULT_FRESHNESS_HOURS))
    
    # Get all news (with caching)
    all_news = _get_all_news()
    
    # Filter by category if specified
    if category_filter != "all":
        all_news = [
            article for article in all_news
            if article["category"] == category_filter
        ]
    else:
        all_news = _prioritize_freshness_by_category(all_news, fresh_hours=fresh_hours)
    
    # Apply limit
    all_news = all_news[:limit]
    
    return jsonify({
        "success": True,
        "count": len(all_news),
        "articles": all_news
    })


def _scrape_article_body(url, timeout=6):
    """Fetch full article text from a URL by extracting <p> tag content."""
    try:
        req = Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; VidyarthiMitra/1.0)"})
        with urlopen(req, timeout=timeout) as resp:
            html = resp.read().decode("utf-8", errors="ignore")
        # Extract all <p> text
        paragraphs = re.findall(r'<p[^>]*>(.*?)</p>', html, re.DOTALL | re.IGNORECASE)
        texts = []
        for p in paragraphs:
            text = re.sub(r'<[^>]+>', '', p).strip()
            if len(text) > 40:  # skip nav/footer snippets
                texts.append(text)
        return " ".join(texts[:30])  # first 30 paragraphs is plenty
    except Exception:
        return ""


@news_bp.route("/news/download")
def download_news():
    """Download all news articles as an Excel file with full scraped body text."""
    import io
    import pandas as pd
    from concurrent.futures import ThreadPoolExecutor, as_completed

    articles = _get_all_news()

    # Scrape full body for each article in parallel (max 15 threads, 6s timeout each)
    bodies = {}
    with ThreadPoolExecutor(max_workers=15) as pool:
        future_map = {pool.submit(_scrape_article_body, a["link"]): i
                      for i, a in enumerate(articles) if a.get("link")}
        for future in as_completed(future_map):
            idx = future_map[future]
            try:
                bodies[idx] = future.result()
            except Exception:
                bodies[idx] = ""

    rows = []
    for i, a in enumerate(articles):
        date_str = a.get("date", "")
        date_only = date_str[:10] if date_str else ""
        time_only = date_str[11:16] if len(date_str) > 10 else ""
        rows.append({
            "Headline":    a.get("title", ""),
            "Date":        date_only,
            "Time":        time_only,
            "Category":    a.get("category", ""),
            "Source":      a.get("source", ""),
            "Summary":     a.get("desc", ""),
            "Full Body":   bodies.get(i, ""),
            "Link":        a.get("link", ""),
        })

    df = pd.DataFrame(rows)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="News")
        ws = writer.sheets["News"]
        col_widths = {"Headline": 55, "Date": 12, "Time": 8, "Category": 18,
                      "Source": 22, "Summary": 50, "Full Body": 100, "Link": 70}
        from openpyxl.styles import Font, PatternFill, Alignment
        for col_cells in ws.columns:
            header = col_cells[0].value
            ws.column_dimensions[col_cells[0].column_letter].width = col_widths.get(header, 20)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="D9252A")
            cell.alignment = Alignment(horizontal="center")
        for row in ws.iter_rows(min_row=2):
            row[5].alignment = Alignment(wrap_text=True)  # Summary
            row[6].alignment = Alignment(wrap_text=True)  # Full Body

    output.seek(0)
    filename = f"vidyarthi-mitra-news-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
